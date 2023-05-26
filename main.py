import math
import openpyxl
import psycorg2

class Station:
    def __init__(self, name, longitude, width, hub, history, year, isSubject, isBMR):
        self.unified = False
        self.name = name
        self.longitude = float(longitude)
        self.width = float(width)
        if hub != None and hub != False:
            self.hub = True
        else:
            self.hub = False
        self.history = history
        if self.history != None:
            self.year = year
        else:
            self.year = False
        if isSubject != None and isSubject!=False:
            self.isSubject = True
        else: self.isSubject = False
        self.list_adjacency = []
        self.isAdjac = False
        self.isBMR = isBMR
    def output (self):
        result = self.name+' '+str(self.longitude)+' '+str(self.width)+' '
        if self.hub:
            result+= 'узловая'
        if self.history!=None:
            result+=' '+self.history
        if self.isBMR:
            result+= ' BMR '
        string_python_shit = ''
        for x in self.list_adjacency:
            string_python_shit = string_python_shit+' '+x.name
        print (result, ':', string_python_shit)
    def isEqual(self, obj2):
        return (self.name == obj2.name and self.longitude == obj2.longitude and self.width == obj2.width and self.hub == obj2.hub)
    def copy(self):
        return Station(self.name, self.longitude, self.width, self.hub, self.history, self.year, self.isSubject, self.isBMR)

class Point:
    def __init__(self, x, y):
        self.x = float(x)
        self.y = float(y)
        self.isHub = False
        self.isDeadEnd = False
        self.history = None
        self.isInVector = False
    def output(self, string_temp):
        print (self.x, self.y, string_temp)

class LineString:
    def __init__(self, init_point, finish_point, history):
        self.init_station = init_point
        self.finish_station = finish_point
        self.list_stations = []
        self.direction = init_point.x<finish_point.x
        self.history = history
        if self.history!=None:
            self.init_station.history = history
            self.finish_station.history = history
        self.isInVector = False
    def length(self):
        return distance(self.init_station, self.finish_station)
    def output(self, string_temp):
        print(self.init_station.x, self.init_station.y, self.finish_station.x, self.finish_station.y, string_temp)

class Vector:
    def __init__(self, list_points):
        self.list_points = list_points
        self.init_station = find_closest_station(list_points[0])
        self.finish_station = find_closest_station(list_points[-1])
        self.list_result = self.history()
    def history(self):
        list_result = []
        for x in self.list_points:
            if x.history!=None:
                list_result.append(x.history)
        return list_result
    def output(self):
        print (self.init_station.name, '-', self.finish_station.name, 'VECTOR')

def distance(point1, point2):
    return ((point1.x-point2.x)**2+(point1.y-point2.y)**2)**0.5

def count_for_point(list_temp, point_temp):
    result = 0
    for x in list_temp:
        if distance(x, point_temp)<0.001:
            result +=1
    #point_temp.output(str(result)+' count_point')
    return result

def create_list_vectors():
    vector_list = []
    for x in hub_list:
        if not x.isInVector:
            x.output('hub')
            temp_linestring = find_first_linestring(x)
            temp_linestring.isInVector = True
            temp_direct = find_other_station_from_line(x, temp_linestring)
            vector_list.append(create_vector(x, temp_direct))
    return vector_list

def create_vector(init_hub, direct_point):
    temp_points_list = []
    temp_points_list.append(init_hub)
    temp_points_list.append(direct_point)
    temp = direct_point
    init_hub.output('init_hub')
    direct_point.output('direct_point')
    init_hub.isInVector = True
    direct_point.isInVector = True
    while not temp.isHub and not temp.isDeadEnd:
        temp.output('temp')
        temp_linestring = find_first_linestring(temp)
        temp_linestring.isInVector = True
        temp = find_other_station_from_line(temp, temp_linestring)
        temp_points_list.append(temp)
        temp.isInVector = True
    print('SUCCESS SUKA')
    if temp.isDeadEnd:
        print('Чурка')
    vector_result = Vector(temp_points_list)
    vector_result.output()
    return vector_result

def find_other_station_from_line(point1, line1):
    if is_point_in_line(point1, line1) == 1:
        return line1.finish_station
    else:
        return line1.init_station

def find_first_linestring(point1):
    for x in main_linestring_list:
        if is_point_in_line(point1, x)>0 and not x.isInVector:
            return x
        #if x.isInVector:
        #    x.output('IsInVector')
    print(point1.x, point1.y, 'ERROR СУКА')
    if point1.isDeadEnd:
        print ('ГОВНО')
    return None

def is_point_in_line(point1, line1):
    if(abs(point1.x-line1.init_station.x)<0.001 and abs(point1.y - line1.init_station.y)<0.001):
        #line1.output('FUCK YOU')
        return 1
    elif(abs(point1.x - line1.finish_station.x)<0.001 and abs(point1.y - line1.finish_station.y)<0.001):
        #line1.output('FUCK YOU')
        return 2
    else:
        return 0

def distance_mixed(point1, station1):
    return distance(point1, Point(station1.longitude, station1.width))

def find_closest_station(point1):
    min_r = 5
    temp = None
    for x in main_list_station:
        temp_r = distance_mixed(point1, x)
        if temp_r<=0:
            print ('ПАРАША')
        if temp_r<min_r:
            min_r = temp_r
            temp = x
    return temp

def excel_save():
    wb = openpyxl.Workbook()
    ws = wb.active
    index = 0
    for x in vector_list:
        temp_l = [x.init_station.name, x.finish_station.name, index]
        ws.append(temp_l)
        index += 1
    for i in range(len(vector_list)):
        ws_temp = wb.create_sheet(title=str(i))
        for x in vector_list[i].list_points:
            temp_l = [x.x, x.y, x.history]
            ws_temp.append(temp_l)
    wb.save('result.xlsx')

if __name__ == '__main__':
    excel_file = openpyxl.load_workbook('finally_ready_station_file.xlsx')
    excel_file = excel_file['Sheet']
    main_list_station = []
    for row in excel_file:
        if row[0].value != 'name' and row[0].value != None:
            main_list_station.append(
                Station(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value,
                        row[6].value, row[7].value))
    linestring_file = openpyxl.load_workbook('linestring_Europe_data.xlsx')
    linestring_file = linestring_file['Sheet']
    main_linestring_list = []
    point_list = []
    index = 0
    for row in linestring_file:
        if row[0].value != 'x_init':
            temp_point_first = Point(row[0].value, row[1].value)
            point_list.append(temp_point_first)
            temp_point_second = Point(row[2].value, row[3].value)
            point_list.append(temp_point_second)
            main_linestring_list.append(
                LineString(temp_point_first, temp_point_second, row[4].value))
            print(index)
            index += 1
    hub_list = []
    dead_end_list = []
    for x in point_list:
        if count_for_point(point_list, x) > 2:
            hub_list.append(x)
            x.isHub = True
        if count_for_point(point_list, x) == 1:
            x.output('DeadEnd')
            dead_end_list.append(x)
            x.isDeadEnd = True
    print (len(dead_end_list), 'Dead_end_list')
    vector_list = create_list_vectors()
    for x in vector_list:
        x.init_station = find_closest_station(x.list_points[0])
        x.finish_station = find_closest_station(x.list_points[-1])
    conn = psycorg2.connect(dbname = 'postgres', user = 'postgres', password = 'Toplova_2006', host = 'localhost')
    cursor = conn.cursor()